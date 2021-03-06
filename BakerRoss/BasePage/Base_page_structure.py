from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
import logging
import csv
from bs4 import BeautifulSoup
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
import openpyxl
import datetime
import sys
import os
import unittest
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from io import StringIO

# Variables
localenv = 'C:\\Users\\Ben\\PycharmProjects'
localdriver = 'C:\\seleniumDriver'
autotestenvironment = localenv + '\\automatedTests\\BakerRoss\\TestsConfiguration\\env.xml'
autotestenvchoice = localenv + '\\automatedTests\\BakerRoss\\TestsConfiguration\\env_choice.csv'
user_catalogue = localenv + '\\automatedTests\\BakerRoss\\TestsData\\user_catalogue.csv'
chromedriver = localdriver + '\\chromedriver_win32\\chromedriver.exe'
firefoxdriver = 'C:\\Program Files (x86)\\Mozilla Firefox\\firefox.exe'
internetexplorerdriver = localdriver + '\\ie_driver\\IEDriverServer.exe'

# Locators
firstBoxPPimgXpath = "//*[@id='page']/div[4]/div/div[4]/div[5]/div[6]/ul[1]/li[1]/a[1]/img"
firstBoxPPlinkXpath = "//*[@id='page']/div[4]/div/div[4]/div[5]/div[6]/ul[1]/li[1]/div[2]/form/a"
megamenufirstlinkXpath = "//*[@id='custommenu-nav']/li[1]/a[1]"

# Load the browser defined in env_choice
with open(autotestenvchoice, "r") as file:
    csv_reader = csv.DictReader(file)
    for line in csv_reader:
        browser = line['browser']
        if browser == 'Chrome':
            driver = webdriver.Chrome(chromedriver)
        if browser == 'Firefox':
            binary = FirefoxBinary(firefoxdriver)
            driver = webdriver.Firefox(firefox_binary=binary)
        if browser == 'IE':
            driver = webdriver.Ie(internetexplorerdriver)

# Actions
def getcountry():
    with open(autotestenvchoice, "r") as file:
        csv_reader = csv.DictReader(file)
        for line in csv_reader:
            country = line['country']
            return country

def impwait():
    driver.implicitly_wait(5)

def createlogfile():
    logfilename = datetime.datetime.now().strftime('%Y-%m-%d-%H.%M.%S') + '_' + os.path.splitext(os.path.basename(sys.argv[0]))[0] + '_log.xlsx'
    wb = openpyxl.Workbook()
    wb.save(logfilename)
    ws = wb.worksheets[0]
    ws['A1'] = 'Date'
    ws['B1'] = 'Test'
    ws['C1'] = 'Status'
    ws['D1'] = 'Details'
    ws['A2'] = datetime.datetime.now().strftime('%Y-%m-%d-%H.%M.%S')
    ws ['B2'] = os.path.splitext(os.path.basename(sys.argv[0]))[0]
    wb.save(logfilename)

def getdomain():
    with open(autotestenvchoice,"r") as file:
        csv_reader = csv.DictReader(file)
        for line in csv_reader:
            environment = line['env']
            country = line['country']
            soup = BeautifulSoup(open(autotestenvironment),'html.parser')
            for a in soup.find_all("env", attrs= {"id" : environment}):
                for b in a.find_all("country", attrs= {"id": country}):
                        domain = b.text
                        return domain

def gotourl(path):
    getdomain()
    url = getdomain() + path
    driver.get(url)
    driver.maximize_window()

def gotoprodurl():
    getdomain()
    url = getdomain()
    driver.get(url)
    driver.maximize_window()
    impwait()
    driver.find_element_by_xpath(megamenufirstlinkXpath).click()
    impwait()
    firstBoxPP = driver.find_element_by_xpath(firstBoxPPimgXpath)
    hover = ActionChains(driver).move_to_element(firstBoxPP)
    hover.perform()
    driver.find_element_by_xpath(firstBoxPPlinkXpath).click()
    impwait()

def getloginvalue():
    site = getcountry()
    with open(user_catalogue, "r") as file:
        csv_reader = csv.DictReader(file)
        row = [row for row in csv_reader if row['country'] == site]
        for rows in row:
            emailValue = rows['login']
            return emailValue

def getpassvalue():
    site = getcountry()
    with open(user_catalogue, "r") as file:
        csv_reader = csv.DictReader(file)
        row = [row for row in csv_reader if row['country'] == site]
        for rows in row:
            passValue = rows['password']
            return passValue