from selenium import webdriver
import csv
from bs4 import BeautifulSoup
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
import openpyxl
import datetime
import sys
import os
from TestsConfiguration.local_env import *
from selenium.webdriver import ActionChains
#import unittest
#from selenium.webdriver.support import expected_conditions as EC
#from selenium.webdriver.common.by import By
#from io import StringIO
#from selenium.webdriver.support.ui import WebDriverWait
#import logging


class BasePage(object):

    # environement and configuration files
    autotestenvironment = localenv + '\\automatedTests\\BakerRoss\\TestsConfiguration\\BakerRoss_env.xml'
    autotestenvchoice = localenv + '\\automatedTests\\BakerRoss\\TestsConfiguration\\env_choice.csv'
    user_catalogue = localenv + '\\automatedTests\\BakerRoss\\TestsData\\user_catalogue.csv'

    # locators
    #firstBoxPPimgXpath = "//*[@id='page']/div[4]/div/div[4]/div[5]/div[6]/ul[1]/li[1]/a[1]/img"
    firstBoxPPimgXpath = "//*[@class='category-products']/ul[1]/li[1]/a[1]/img"
    firstBoxPPlinkXpath = "//*[@id='page']/div[4]/div/div[4]/div[5]/div[6]/ul[1]/li[1]/div[2]/form/a"
    megamenufirstlinkXpath = "//*[@id='custommenu-nav']/li[3]/a[1]"

    # Load the browser defined in env_choice
    with open(autotestenvchoice, "r") as file:
        csv_reader = csv.DictReader(file)
        for line in csv_reader:
            browser = line['browser']
            if browser == 'Chrome':
                driver = webdriver.Chrome(chromedriver)
            if browser == 'Firefox':
                binary = FirefoxBinary(firefoxdriver)
                driver = webdriver.Firefox(firefox_binary=binary)
            if browser == 'IE':
                driver = webdriver.Ie(internetexplorerdriver)

    # Methods
    def __init__(self, autotestenvironment, autotestenvchoice, user_catalogue, driver):
        # Variables
        self.autotestenvironment = autotestenvironment
        self.autotestenvchoice = autotestenvchoice
        self.user_catalogue = user_catalogue
        self.driver = driver

    def getcountry(self):
        with open(self.autotestenvchoice, "r") as file:
            csv_reader = csv.DictReader(file)
            for line in csv_reader:
                country = line['country']
                return country

    def impwait(self):
        self.driver.implicitly_wait(5)

    def createlogfile(self):
        logfilename = datetime.datetime.now().strftime('%Y-%m-%d-%H.%M.%S') + '_' + os.path.splitext(os.path.basename(sys.argv[0]))[0] + '_log.xlsx'
        wb = openpyxl.Workbook()
        wb.save(logfilename)
        ws = wb.worksheets[0]
        ws['A1'] = 'Date'
        ws['B1'] = 'Test'
        ws['C1'] = 'Status'
        ws['D1'] = 'Details'
        ws['A2'] = datetime.datetime.now().strftime('%Y-%m-%d-%H.%M.%S')
        ws ['B2'] = os.path.splitext(os.path.basename(sys.argv[0]))[0]
        wb.save(logfilename)

    def getdomain(self):
        with open(self.autotestenvchoice,"r") as file:
            csv_reader = csv.DictReader(file)
            for line in csv_reader:
                environment = line['env']
                country = line['country']
                soup = BeautifulSoup(open(self.autotestenvironment),'html.parser')
                for a in soup.find_all("env", attrs={"id": environment}):
                    for b in a.find_all("country", attrs={"id": country}):
                            domain = b.text
                            return domain

    def gotourl(self, path):
        url = self.getdomain() + path
        self.driver.get(url)
        self.driver.maximize_window()
        self.impwait()

    def gotolistingpagereview(self, pathwithreview):
        url = self.getdomain() + pathwithreview
        self.driver.get(url)
        self.driver.maximize_window()
        self.impwait()

    def gotoprodurl(self):
        url = self.getdomain()
        self.driver.get(url)
        self.driver.maximize_window()
        self.impwait()
        self.driver.find_element_by_xpath(self.megamenufirstlinkXpath).click()
        self.impwait()
        firstBoxPP = self.driver.find_element_by_xpath(self.firstBoxPPimgXpath)
        hover = ActionChains(self.driver).move_to_element(firstBoxPP)
        hover.perform()
        self.driver.find_element_by_xpath(self.firstBoxPPlinkXpath).click()
        self.impwait()

    def gotolistingpage(self):
        self.getdomain()
        url = self.getdomain()
        self.driver.get(url)
        self.driver.maximize_window()
        self.impwait()
        self.driver.find_element_by_xpath(self.megamenufirstlinkXpath).click()
        self.impwait()

    def getloginvalue(self):
        site = self.getcountry()
        with open(self.user_catalogue, "r") as file:
            csv_reader = csv.DictReader(file)
            row = [row for row in csv_reader if row['country'] == site]
            for rows in row:
                emailValue = rows['login']
                return emailValue

    def getpassvalue(self):
        site = self.getcountry()
        with open(self.user_catalogue, "r") as file:
            csv_reader = csv.DictReader(file)
            row = [row for row in csv_reader if row['country'] == site]
            for rows in row:
                passValue = rows['password']
                return passValue